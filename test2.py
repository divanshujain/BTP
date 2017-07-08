from pygame import mixer
import time
from openpyxl import Workbook, load_workbook

class _GetchUnix:
    def __init__(self):
        import tty, sys

    def __call__(self):
        import sys, tty, termios
        fd = sys.stdin.fileno()
        old_settings = termios.tcgetattr(fd)
        try:
            tty.setraw(sys.stdin.fileno())
            ch = sys.stdin.read(1)
        finally:
            termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)
        return ch


getch = _GetchUnix()

ans = [['bleh', 'Up', 'Down', 'Same', 'Up'], ['bleh', 'Up', 'Up', 'Down', 'Same', 'Down'], ['bleh', 'Up', 'Up', 'Down', 'Down', 'Same']]
wb = load_workbook('test.xlsx')
ws = wb.active
Test_number = input("Test Number : ")
final = 0
def startTest(File, D):
    global final
    c = 0
    add = []
    if D == 0:
        T = xrange(1,5)
    else:
        T = xrange(1,6)
    print(D)
    for a in T:
        print("Press Enter to move to next set\n")
        key = ord(getch())
        mixer.init()
        mixer.music.load('./Sounds2/ramptest'+str(D+1)+str(a)+'.ogg')
        mixer.music.play()
        time.sleep(8)
        print("Final Tone")
        while mixer.music.get_busy():
            key = ord(getch())
            if key == 27: #ESC
                mixer.music.stop()
                break
            elif key == 52:
                A = "Down"
                break
            elif key == 54:
                A = "Up"
                break
            elif key == 53:
                A = "Same"
                break

        if A == ans[D][a]:
            c+=1
            print("Correct\n")
            add.append('Correct')
            File.write("Correct\n")
        else:
            print("Wrong")
            print(ans[D][a]+'\n')
            add.append('Wrong Correct:'+ans[D][a]+' Inputted:'+A)
            File.write("Wrong Correct:"+ans[D][a]+" Inputted:"+A+"\n")
    File.write("Total Correct : "+str(c)+"\n")
    if D>0:
        if D == 1:
            final += (8*c)
        else:
            final += (12*c)

        ws.cell(row=(D*6)+1, column=Test_number).value = c
        for i in xrange(1,6):
            t = (D*6)+1+i
            ws.cell(row=t, column=Test_number).value = add[i-1]
    return


Name = raw_input("Enter name : ")
Age = input("Enter Age : ")
Gender = raw_input("Gender (M/F/NB) : ")
Med = raw_input("Any hearing disablities : ")
Train = raw_input("Musical training (None/Mod/Adv) : ")
Fam = raw_input("Family members with musical training (Y/N) : ")
f = open(Name+'.txt',"w+")
f.write("Age : "+str(Age)+"\n")
f.write("Gender : "+Gender+"\n")
f.write("Medical Conditions : "+Med+"\n")
f.write("Musical Training : "+Train+"\n")
f.write("Family BG : "+Fam+"\n")

ws.cell(row=1, column=Test_number).value = 'Name : '+Name
ws.cell(row=2, column=Test_number).value = 'Age : '+str(Age)
ws.cell(row=3, column=Test_number).value = 'Gender : '+Gender
ws.cell(row=4, column=Test_number).value = 'Medical Conditions : '+Med
ws.cell(row=5, column=Test_number).value = 'Musical Training : '+Train
ws.cell(row=6, column=Test_number).value = 'Family BG : '+Fam

for d in range(0,3):
    print("Press Enter to start Test Or Escape to Exit\n")
    key = ord(getch())
    if key == 27: #ESC
        break
    elif key == 13: #Enter
        f.write(str(d)+'\n')
        startTest(f, d)

ws.cell(row=19, column=Test_number).value = final
wb.save('test.xlsx')
f.close()
